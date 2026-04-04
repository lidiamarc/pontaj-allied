import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
from supabase import create_client
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Pontaj Allied Engineers", page_icon="🏗️", layout="wide")

# ============================================================
# SUPABASE CONNECTION
# ============================================================
@st.cache_resource
def get_supabase():
    url = "https://rhvaumovrmgljlhjnscl.supabase.co"
    key = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InJodmF1bW92cm1nbGpsaGpuc2NsIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzUyNDE5MTMsImV4cCI6MjA5MDgxNzkxM30.xXnXljiRES2AdrnwJKv4WOUwtQZGrBZwG7aKVLPtVNA"
    return create_client(url, key)

supabase = get_supabase()

# ============================================================
# SARBATORI LEGALE ROMANIA 2026
# ============================================================
SARBATORI_ROMANIA = [
    date(2026, 1, 1), date(2026, 1, 2), date(2026, 4, 10), date(2026, 4, 12),
    date(2026, 4, 13), date(2026, 5, 1), date(2026, 6, 1), date(2026, 6, 8),
    date(2026, 6, 9), date(2026, 8, 15), date(2026, 11, 30),
    date(2026, 12, 1), date(2026, 12, 25), date(2026, 12, 26),
]

CONFIG_DEFAULT = {
    "proiecte": [
        "907", "000.CO", "010.INVATARE", "020.DIVERSE", "03.PROIECTE LICITATIE",
        "111/112", "114.OCT-C1", "115.OCT-C2", "1305.ARENA NATIONALA",
        "132.SNM", "133.SNM", "136.BORZESTI", "137.COROD", "138.LEMACONS",
        "140.WNDPT", "141.PESTERA2", "142.POIANA", "145.TOPALU", "144.CNV",
        "143.MIRCEA", "191.CEF-BBD", "192.CEF-PTR", "193.CEF-GNS1",
        "194.CEF-GNS2", "195.CEF-PLC", "210.H4L", "227A.DECO BHB",
        "490.ELIADE", "575.TN2", "675.OLD", "678.OCP2", "680.PENINSULA",
        "691.OLC", "700.OHD", "705.OMN", "706.J8", "722.WNDGP", "723.WNDGC",
        "724.WNDIN", "726.WNDGV", "727.WNDPT", "747.SRUCJ IASI",
        "750.HALA FORD", "752.SPITAL CLUJ", "753.AZUGA", "756.PENNY",
        "768.BUC-3", "777.DOBROSLOVENI", "778.SP-PITESTI", "782.OCCIDENTULUI",
        "797.OST", "799.BCV", "800.AFI3", "803.CHTL", "817.GLOBE",
        "824.TANDAREI", "824.CEF Tandarei", "829.BUC-3", "830.OP17",
        "836.WNDGN", "840.SLT", "842.WNDGP", "853.RMW", "859.SATU MARE",
        "869.KFLBVT", "870.OPL", "872.BANEASA GIURGIU", "874.WNDAO",
        "875.HOLBAN", "878.BC SLT", "881.GHE", "885.AHC", "890.SINAIA",
        "891.DM-BRATIANU", "892.KFLOTP", "900.QDN", "905.01 KFLISD",
        "905.02 KFLSCL", "905.03 KFLIST", "905.04 KFLVLT", "905.KFLLGP",
        "915.SPITAL CONSTANTA", "920.JND", "921.CNV", "923.NOVO PARK",
        "929.BUCHAREST ONE", "936.MORA", "939.ORION DEBRECEN", "941.EDT",
        "943.ARENA NATIONALA", "947.HJLV", "950.KFLVLT", "952.CALIMANESTI",
        "953.Gura Ialomitei", "956.Brazi", "963.KFLJLV", "964.HOLBAN",
        "966.STC", "967.BUC4", "969.GALA", "972.BESS IS3", "AVALANSEI",
        "CASA - VALULUI", "DIVERSE", "FUNDENI", "GW Tower", "OME",
        "renovatio", "SPITAL BACAU", "EVERGENT", "CEE POIANA",
        "H4L.HOLBAN", "H4L.POMPEIU"
    ],
    "colegi": [
        "ADINA-MADALINA DOSPINA", "ADRIANA STEFAN", "ALEXANDRU SETREANU",
        "ALIN MIHAI MINDOIU", "ANAS ALISSA", "ANAMARIA LAZAROIU",
        "ANCA CERCEL", "ANDREI CATRINOIU", "ANDREI PAUN",
        "ANDREI-DANIEL MOROITA", "ARINA-LAURA MITU", "BIANCA ANDRONACHE",
        "BIANCA GABRIELA MANOLACHE", "CEZARINA FLORINA BRAN", "CRISTI ANDONE",
        "DIANA PANA", "DRAGOS GUBICS", "DRAGOS-NICOLAE SEITAN",
        "ELENA BURLACU", "ELENA VOLCINSCHI", "FLORIAN DANIEL TILIMPEA",
        "GABRIELA IONITA", "GABRIELA MARIN", "GEORGIANA ANA-MARIA BALUTA",
        "IOANA GAVRILA", "IONUT-FLORIN TOMA", "IRINA FARCAS",
        "IRINA-MIHAELA MATEI", "LEONARD IONUT TRAZNEA", "LIDIA DRAGANOVA",
        "LIDIA MIHAELA PANTIRU-MARIN", "MADALINA MUSAT", "MIHAI NEGRITORU",
        "OCTAVIAN DEACONU", "RADU SECATUREANU", "RADU STANCIU",
        "RICHARD MILOSAN", "RUXANDRA MARIA ANDRITOIU", "SAMUEL TOMA",
        "STEFAN GEORGESCU", "TIBERIU TUGURLAN", "TUDOR CIUCAN",
        "VIRGIL MICLIUC", "VLAD ANDREI WEBER"
    ],
    "faze": [
        "ASISTENTA TEHNICA", "PT", "DE", "CONCEPT", "CONCEPT PRELIMINAR",
        "CONCEPT FINAL", "DTAC", "EVALUARE OFERTA", "VERIFICARE PROIECT",
        "PTh", "LUCRARI SUPLIMENTARE"
    ],
    "subfaze": [
        "ESTIMARE LISTA DE CANTITATI", "întocmire documentație",
        "ACTUALIZARE MODEL", "Calcul piloti", "Coordonare echipa",
        "Editare memoriu", "EXPERTIZA TEHNICA", "PUNCT DE VEDERE", "REAUTORIZARE"
    ]
}

# ============================================================
# STORAGE FUNCTIONS - SUPABASE
# ============================================================
def incarca_pontaj():
    try:
        result = supabase.table("pontaj").select("*").execute()
        if result.data:
            df = pd.DataFrame(result.data)
            return df
        return pd.DataFrame(columns=["id","data","angajat","proiect","faza","ore","observatii","created_at","coleg","saptamana","data_zi","subfaza","comentarii","timestamp"])
    except Exception as e:
        st.error(f"Eroare la încărcarea datelor: {e}")
        return pd.DataFrame()

def salveaza_inregistrare(inreg):
    try:
        data = {
            "data": inreg.get("data_zi"),
            "angajat": inreg.get("coleg"),
            "proiect": inreg.get("proiect"),
            "faza": inreg.get("faza"),
            "ore": float(inreg.get("ore", 0)),
            "observatii": inreg.get("comentarii", ""),
            "coleg": inreg.get("coleg"),
            "saptamana": inreg.get("saptamana"),
            "data_zi": inreg.get("data_zi"),
            "subfaza": inreg.get("subfaza", ""),
            "comentarii": inreg.get("comentarii", ""),
            "timestamp": inreg.get("timestamp"),
        }
        supabase.table("pontaj").insert(data).execute()
        return True
    except Exception as e:
        st.error(f"Eroare la salvare: {e}")
        return False

def incarca_config():
    if "pontaj_config_v1" not in st.session_state:
        st.session_state["pontaj_config_v1"] = CONFIG_DEFAULT.copy()
    return st.session_state["pontaj_config_v1"]

def salveaza_config(config):
    st.session_state["pontaj_config_v1"] = config
    return True

# ============================================================
# EXPORT GOOGLE SHEETS (XLSX)
# ============================================================
def genereaza_xlsx(df_export, titlu="Pontaj Allied Engineers"):
    wb = Workbook()

    # ---- Sheet 1: Date complete ----
    ws1 = wb.active
    ws1.title = "Date Pontaj"

    # Culori
    culoare_header = "1F4E79"      # albastru inchis
    culoare_header2 = "2E75B6"     # albastru mediu
    culoare_rand_par = "DEEAF1"    # albastru foarte deschis
    culoare_total = "FCE4D6"       # portocaliu deschis pentru totaluri

    thin = Side(style='thin', color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Titlu
    ws1.merge_cells("A1:I1")
    ws1["A1"] = titlu
    ws1["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws1["A1"].fill = PatternFill("solid", start_color=culoare_header)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    # Subtitlu data export
    ws1.merge_cells("A2:I2")
    ws1["A2"] = f"Export generat: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws1["A2"].font = Font(name="Arial", italic=True, size=10, color="595959")
    ws1["A2"].alignment = Alignment(horizontal="center")
    ws1.row_dimensions[2].height = 18

    # Rand gol
    ws1.row_dimensions[3].height = 8

    # Coloane de afisat
    coloane = ["data_zi", "saptamana", "coleg", "proiect", "faza", "subfaza", "ore", "comentarii", "timestamp"]
    headere = ["Data", "Săptămâna", "Coleg", "Proiect", "Faza", "Subfaza", "Ore", "Comentarii", "Înregistrat la"]

    # Header tabel
    for col_idx, header in enumerate(headere, 1):
        cell = ws1.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=culoare_header2)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws1.row_dimensions[4].height = 22

    # Date
    df_clean = df_export.copy()
    for col in coloane:
        if col not in df_clean.columns:
            df_clean[col] = ""

    for row_idx, row in enumerate(df_clean[coloane].itertuples(index=False), 5):
        fill_color = culoare_rand_par if row_idx % 2 == 0 else "FFFFFF"
        for col_idx, val in enumerate(row, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=fill_color)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if col_idx == 7:  # Ore - centrat
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Rand TOTAL
    total_row = len(df_clean) + 5
    ws1.cell(row=total_row, column=6, value="TOTAL ORE:").font = Font(name="Arial", bold=True, size=11)
    ws1.cell(row=total_row, column=6).alignment = Alignment(horizontal="right")
    total_formula = f"=SUM(G5:G{total_row - 1})"
    total_cell = ws1.cell(row=total_row, column=7, value=total_formula)
    total_cell.font = Font(name="Arial", bold=True, size=12, color="C00000")
    total_cell.fill = PatternFill("solid", start_color=culoare_total)
    total_cell.alignment = Alignment(horizontal="center", vertical="center")
    total_cell.border = border
    ws1.row_dimensions[total_row].height = 22

    # Latimi coloane
    latimi = [12, 14, 30, 22, 22, 22, 8, 30, 18]
    for i, latime in enumerate(latimi, 1):
        ws1.column_dimensions[get_column_letter(i)].width = latime

    # ---- Sheet 2: Rezumat pe colegi ----
    ws2 = wb.create_sheet("Rezumat Colegi")
    ws2["A1"] = "Rezumat ore per coleg"
    ws2["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", start_color=culoare_header)
    ws2.merge_cells("A1:C1")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    for col_idx, h in enumerate(["Coleg", "Total Ore", "Nr. Înregistrări"], 1):
        cell = ws2.cell(row=2, column=col_idx, value=h)
        cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=culoare_header2)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    if "coleg" in df_clean.columns and "ore" in df_clean.columns:
        df_clean["ore"] = pd.to_numeric(df_clean["ore"], errors="coerce")
        rezumat_colegi = df_clean.groupby("coleg").agg(
            total_ore=("ore", "sum"),
            nr_inreg=("ore", "count")
        ).sort_values("total_ore", ascending=False).reset_index()

        for row_idx, row in rezumat_colegi.iterrows():
            fill_color = culoare_rand_par if row_idx % 2 == 0 else "FFFFFF"
            ws2.cell(row=row_idx+3, column=1, value=row["coleg"]).font = Font(name="Arial", size=10)
            ws2.cell(row=row_idx+3, column=1).fill = PatternFill("solid", start_color=fill_color)
            ws2.cell(row=row_idx+3, column=1).border = border
            ore_cell = ws2.cell(row=row_idx+3, column=2, value=round(row["total_ore"], 1))
            ore_cell.font = Font(name="Arial", size=10)
            ore_cell.fill = PatternFill("solid", start_color=fill_color)
            ore_cell.alignment = Alignment(horizontal="center")
            ore_cell.border = border
            nr_cell = ws2.cell(row=row_idx+3, column=3, value=int(row["nr_inreg"]))
            nr_cell.font = Font(name="Arial", size=10)
            nr_cell.fill = PatternFill("solid", start_color=fill_color)
            nr_cell.alignment = Alignment(horizontal="center")
            nr_cell.border = border

    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 18

    # ---- Sheet 3: Rezumat pe proiecte ----
    ws3 = wb.create_sheet("Rezumat Proiecte")
    ws3["A1"] = "Rezumat ore per proiect"
    ws3["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws3["A1"].fill = PatternFill("solid", start_color=culoare_header)
    ws3.merge_cells("A1:C1")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 26

    for col_idx, h in enumerate(["Proiect", "Total Ore", "% din Total"], 1):
        cell = ws3.cell(row=2, column=col_idx, value=h)
        cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=culoare_header2)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    if "proiect" in df_clean.columns and "ore" in df_clean.columns:
        df_clean["ore"] = pd.to_numeric(df_clean["ore"], errors="coerce")
        rezumat_proiecte = df_clean.groupby("proiect")["ore"].sum().sort_values(ascending=False).reset_index()
        total_ore = rezumat_proiecte["ore"].sum()

        for row_idx, row in rezumat_proiecte.iterrows():
            fill_color = culoare_rand_par if row_idx % 2 == 0 else "FFFFFF"
            ws3.cell(row=row_idx+3, column=1, value=row["proiect"]).font = Font(name="Arial", size=10)
            ws3.cell(row=row_idx+3, column=1).fill = PatternFill("solid", start_color=fill_color)
            ws3.cell(row=row_idx+3, column=1).border = border
            ore_cell = ws3.cell(row=row_idx+3, column=2, value=round(row["ore"], 1))
            ore_cell.font = Font(name="Arial", size=10)
            ore_cell.fill = PatternFill("solid", start_color=fill_color)
            ore_cell.alignment = Alignment(horizontal="center")
            ore_cell.border = border
            pct = round(row["ore"] / total_ore * 100, 1) if total_ore > 0 else 0
            pct_cell = ws3.cell(row=row_idx+3, column=3, value=f"{pct}%")
            pct_cell.font = Font(name="Arial", size=10)
            pct_cell.fill = PatternFill("solid", start_color=fill_color)
            pct_cell.alignment = Alignment(horizontal="center")
            pct_cell.border = border

    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 14

    # Salveaza in memorie
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ============================================================
# UTILITARE
# ============================================================
def get_saptamani_2026():
    saptamani = []
    vazute = set()
    d = date(2026, 1, 1)
    while d.year == 2026:
        luni = d - timedelta(days=d.weekday())
        duminica = luni + timedelta(days=6)
        week_num = d.isocalendar()[1]
        label = f"W{week_num:02d} — {luni.strftime('%d.%m')} – {duminica.strftime('%d.%m.%Y')}"
        if label not in vazute:
            saptamani.append((label, f"w{week_num:02d}", luni, duminica))
            vazute.add(label)
        d += timedelta(days=7)
    return saptamani

def get_zile_saptamana(luni, duminica):
    zile = []
    d = luni
    while d <= duminica:
        e_sarbatoare = d in SARBATORI_ROMANIA
        e_weekend = d.weekday() >= 5
        nume_zi = ["Luni", "Marți", "Miercuri", "Joi", "Vineri", "Sâmbătă", "Duminică"][d.weekday()]
        label = f"{nume_zi} {d.strftime('%d.%m')}"
        if e_sarbatoare: label += " 🔴"
        elif e_weekend: label += " ⚫"
        zile.append((label, d, e_weekend or e_sarbatoare))
        d += timedelta(days=1)
    return zile

# ============================================================
# APP
# ============================================================
config = incarca_config()

with st.sidebar:
    st.markdown("### 🏗️ Allied Engineers")
    st.markdown("---")
    pagina = st.radio("Navigare", ["📝 Introducere ore", "📊 Rapoarte", "⚙️ Administrare"], label_visibility="collapsed")
    st.markdown("---")
    st.caption("Allied Engineers MEP © 2026")

# ============================================================
# PAGINA 1 - INTRODUCERE ORE
# ============================================================
if pagina == "📝 Introducere ore":
    st.title("📝 Pontaj ore")
    st.markdown("Completează orele lucrate pentru săptămâna selectată.")
    st.markdown("---")

    col1, col2 = st.columns(2)
    with col1:
        coleg = st.selectbox("👤 Numele tău", ["— selectează —"] + sorted(config["colegi"]))
        saptamani = get_saptamani_2026()
        azi = date.today()
        sapt_idx = 0
        for i, (label, cod, luni, dum) in enumerate(saptamani):
            if luni <= azi <= dum:
                sapt_idx = i
                break
        sapt_label = st.selectbox("📅 Săptămâna", [s[0] for s in saptamani], index=sapt_idx)
        sapt_info = next(s for s in saptamani if s[0] == sapt_label)
        sapt_cod, sapt_luni, sapt_dum = sapt_info[1], sapt_info[2], sapt_info[3]

    with col2:
        st.markdown("**📆 Zilele săptămânii:**")
        zile = get_zile_saptamana(sapt_luni, sapt_dum)
        for lz, dz, el in zile:
            st.caption(f"~~{lz}~~" if el else lz)

    st.markdown("---")

    if coleg == "— selectează —":
        st.info("👆 Selectează-ți numele pentru a continua.")
    else:
        st.subheader(f"Adaugă intrare pentru {coleg}")
        with st.form("form_pontaj", clear_on_submit=True):
            ca, cb, cc = st.columns(3)
            with ca:
                zile_lucr = [(l, d) for l, d, el in zile if not el]
                zi_label = st.selectbox("📅 Ziua", [l for l, d in zile_lucr])
                zi_data = next(d for l, d in zile_lucr if l == zi_label)
                ore = st.number_input("⏱️ Ore lucrate", min_value=0.5, max_value=24.0, value=8.0, step=0.5)
            with cb:
                proiect = st.selectbox("🏗️ Proiect", sorted(config["proiecte"]))
                faza = st.selectbox("📐 Faza", config["faze"])
            with cc:
                sf_opt = ["— fără subfază —"] + config["subfaze"]
                subfaza = st.selectbox("🔍 Subfază", sf_opt)
                if subfaza == "— fără subfază —": subfaza = ""
                comentarii = st.text_input("💬 Comentarii", placeholder="opțional...")

            if st.form_submit_button("✅ Salvează înregistrarea", use_container_width=True):
                inreg = {
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "coleg": coleg, "saptamana": sapt_cod,
                    "data_zi": zi_data.strftime("%Y-%m-%d"),
                    "proiect": proiect, "faza": faza, "subfaza": subfaza,
                    "ore": ore, "comentarii": comentarii
                }
                if salveaza_inregistrare(inreg):
                    st.success(f"✅ Salvat! {ore}h pe {proiect} — {faza}")
                    st.rerun()

        st.markdown("---")
        st.subheader(f"Intrările tale din {sapt_label.split('—')[0].strip()}")
        df = incarca_pontaj()
        if len(df) > 0:
            df["ore"] = pd.to_numeric(df["ore"], errors="coerce")
            df_cs = df[(df["coleg"] == coleg) & (df["saptamana"] == sapt_cod)]
            if len(df_cs) == 0:
                st.info("Nu ai înregistrări pentru această săptămână.")
            else:
                st.metric("Total ore săptămână", f"{df_cs['ore'].sum():.1f}h")
                d2 = df_cs[["data_zi", "proiect", "faza", "subfaza", "ore", "comentarii"]].copy()
                d2.columns = ["Data", "Proiect", "Faza", "Subfaza", "Ore", "Comentarii"]
                d2["Data"] = pd.to_datetime(d2["Data"]).dt.strftime("%d.%m")
                st.dataframe(d2, use_container_width=True, hide_index=True)
        else:
            st.info("Nu ai înregistrări pentru această săptămână.")

# ============================================================
# PAGINA 2 - RAPOARTE
# ============================================================
elif pagina == "📊 Rapoarte":
    st.title("📊 Rapoarte pontaj")
    st.markdown("---")
    df = incarca_pontaj()
    if len(df) == 0:
        st.info("Nu există date de pontaj încă.")
    else:
        df["ore"] = pd.to_numeric(df["ore"], errors="coerce")
        c1, c2, c3 = st.columns(3)
        with c1:
            sd = sorted(df["saptamana"].dropna().unique().tolist())
            sf = st.multiselect("Săptămâna", sd, default=sd[-1:] if sd else [])
        with c2:
            cf = st.multiselect("Coleg", sorted(df["coleg"].dropna().unique().tolist()))
        with c3:
            pf = st.multiselect("Proiect", sorted(df["proiect"].dropna().unique().tolist()))

        df_f = df.copy()
        if sf: df_f = df_f[df_f["saptamana"].isin(sf)]
        if cf: df_f = df_f[df_f["coleg"].isin(cf)]
        if pf: df_f = df_f[df_f["proiect"].isin(pf)]

        st.markdown("---")
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total ore", f"{df_f['ore'].sum():.0f}h")
        m2.metric("Colegi activi", df_f["coleg"].nunique())
        m3.metric("Proiecte active", df_f["proiect"].nunique())
        m4.metric("Înregistrări", len(df_f))
        st.markdown("---")

        t1, t2, t3 = st.tabs(["Per coleg", "Per proiect", "Per fază"])
        with t1:
            r = df_f.groupby("coleg")["ore"].sum().sort_values(ascending=False).reset_index()
            r.columns = ["Coleg", "Ore"]; r["Ore"] = r["Ore"].round(1)
            st.dataframe(r, use_container_width=True, hide_index=True)
        with t2:
            r = df_f.groupby("proiect")["ore"].sum().sort_values(ascending=False).reset_index()
            r.columns = ["Proiect", "Ore"]; r["Ore"] = r["Ore"].round(1)
            t = r["Ore"].sum()
            r["Procent"] = (r["Ore"] / t * 100).round(1).astype(str) + "%"
            st.dataframe(r, use_container_width=True, hide_index=True)
        with t3:
            r = df_f.groupby("faza")["ore"].sum().sort_values(ascending=False).reset_index()
            r.columns = ["Faza", "Ore"]; r["Ore"] = r["Ore"].round(1)
            st.dataframe(r, use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("### 📥 Export pentru Google Sheets")
        col_exp1, col_exp2 = st.columns(2)

        with col_exp1:
            xlsx_data = genereaza_xlsx(df_f, "Pontaj Allied Engineers")
            st.download_button(
                label="📊 Descarcă Excel / Google Sheets",
                data=xlsx_data,
                file_name=f"pontaj_allied_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                help="Descarcă fișier .xlsx cu 3 foi: date complete, rezumat colegi, rezumat proiecte"
            )

        with col_exp2:
            st.info("💡 **Cum încarci în Google Sheets:**\n1. Descarcă fișierul\n2. Mergi pe sheets.google.com\n3. Fișier → Importă → alegi fișierul\n4. Gata!")

# ============================================================
# PAGINA 3 - ADMINISTRARE
# ============================================================
elif pagina == "⚙️ Administrare":
    st.title("⚙️ Administrare liste")
    parola = st.text_input("🔒 Parolă admin", type="password")
    if parola != "allied2026":
        if parola: st.error("Parolă incorectă.")
        st.info("Introdu parola pentru a accesa administrarea.")
        st.stop()

    st.success("✅ Acces acordat.")
    st.markdown("---")
    t1, t2, t3, t4 = st.tabs(["Proiecte", "Colegi", "Faze", "Subfaze"])

    def editor(tab, cheie, titlu):
        with tab:
            st.subheader(titlu)
            nou = st.text_area("Un element pe linie:", value="\n".join(config[cheie]), height=400, key=f"ed_{cheie}")
            if st.button(f"💾 Salvează {titlu}", key=f"sv_{cheie}"):
                elemente = [x.strip() for x in nou.split("\n") if x.strip()]
                config[cheie] = elemente
                salveaza_config(config)
                st.success(f"✅ {titlu} actualizate! ({len(elemente)} elemente)")
                st.rerun()

    editor(t1, "proiecte", "Proiecte")
    editor(t2, "colegi", "Colegi")
    editor(t3, "faze", "Faze")
    editor(t4, "subfaze", "Subfaze")

    st.markdown("---")
    st.subheader("🗃️ Export complet date")
    df = incarca_pontaj()
    if len(df) > 0:
        st.write(f"Total înregistrări: **{len(df)}**")
        st.dataframe(df, use_container_width=True)

        st.markdown("### 📥 Export pentru Google Sheets")
        col_exp1, col_exp2 = st.columns(2)

        with col_exp1:
            xlsx_data = genereaza_xlsx(df, "Pontaj Allied Engineers — COMPLET")
            st.download_button(
                label="📊 Descarcă TOATE datele — Excel / Google Sheets",
                data=xlsx_data,
                file_name=f"pontaj_complet_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        with col_exp2:
            st.info("💡 **Cum încarci în Google Sheets:**\n1. Descarcă fișierul\n2. Mergi pe sheets.google.com\n3. Fișier → Importă → alegi fișierul\n4. Gata!")
    else:
        st.info("Nu există date salvate încă.")
